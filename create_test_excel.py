import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, time, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Headers (Dates) in Row 1
base_date = datetime.now().date()
# Start from next Monday
days_ahead = 0 - base_date.weekday()
if days_ahead <= 0: # Target next Monday
    days_ahead += 7
next_monday = base_date + timedelta(days=days_ahead)

dates = [next_monday + timedelta(days=i) for i in range(5)]
for col, d in enumerate(dates, start=2):
    ws.cell(row=1, column=col, value=d)

# Times in Column A
times = [time(hour=h, minute=0) for h in range(8, 18)]
for row, t in enumerate(times, start=2):
    ws.cell(row=row, column=1, value=t)

# Add a "Blue" cell for PHYS2002
# Let's say Monday at 10:00 AM (Row 4, Col 2)
# And Wednesday at 2:00 PM (Row 8, Col 4)

blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

# Slot 1
cell1 = ws.cell(row=4, column=2, value="PHYS2002")
cell1.fill = blue_fill

# Slot 2
cell2 = ws.cell(row=8, column=4, value="PHYS2002")
cell2.fill = blue_fill

wb.save("schedule.xlsx")
print("schedule.xlsx created with dummy data.")
