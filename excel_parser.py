import openpyxl
from datetime import datetime, timedelta, time
import config

def parse_schedule(file_path):
    """
    Parses the Excel file and extracts time slots from cells with specific background colors.
    Returns a list of dictionaries: {'start': datetime, 'end': datetime, 'title': str}
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[config.SHEET_NAME]
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return []

    slots = []
    
    # Iterate through rows and columns to find colored cells
    # This is a heuristic approach; needs refinement based on actual sheet structure
    # Assumption: The sheet has dates in headers or rows, and times in the other axis.
    # For now, let's assume a simple list format or grid where blue cells indicate slots.
    
    # Simplest valid approach without seeing the file:
    # Iterate all cells, if blue, try to infer date/time from row/col headers. 
    # Since I don't know the layout, I will implement a generic "find blue cells" and 
    # assume the cell CONTENT describes the slot or the headers do.
    
    # BETTER APPROACH for "PHYS2002 slots":
    # Maybe the text in the cell says "PHYS2002"? 
    # The prompt says "relevant PHYS2002 slots in blue".
    
    print("Scanning for blue cells...")
    for row in sheet.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.start_color:
                color = cell.fill.start_color
                # default implementation for simple solid fills
                # openpyxl colors can be '00RRGGBB' or theme colors.
                
                color_hex = "None"
                if color.type == 'rgb':
                    color_hex = color.rgb # e.g. "FF0000FF"
                elif color.type == 'indexed':
                    # Indexed colors are harder, usually legacy.
                    pass
                elif color.type == 'theme':
                    # Theme colors require looking up the theme (complex).
                    # Often the user sees "Blue" but it's Theme 4, tint 0.2.
                    # For now, we log what we find if debugging.
                    pass 
                
                # Check if it matches our target (stripping alpha if present)
                # target hexes in config are just examples.
                
                # Heuristic: if the user said "PHYS2002 slots in blue", 
                # let's also check if "PHYS2002" is in the text if it exists.
                cell_text = str(cell.value) if cell.value else ""
                
                # If we verify it is the right slot, we need to extract Time and Date.
                # Let's assume for this implementation that row 1 has Dates and Col A has Times.
                # This is a standard schedule layout.
                
                extracted_slot = extract_slot_from_cell(sheet, cell)
                if extracted_slot:
                     # Check color matches OR text matches if we can't rely on exact color code yet
                    if "PHYS2002" in cell_text or is_blue(color):
                        slots.append(extracted_slot)

    return slots

def is_blue(color_obj):
    # This needs to be robust. 
    # For now, just return True to verify logic flow if text matches, 
    # or check against the config list.
    if color_obj.rgb in config.TARGET_COLOR_HEXES:
        return True
    return False

def extract_slot_from_cell(sheet, cell):
    # Try to find Date from Row 1, Time from Col A
    # cell.row is 1-based, cell.column is 1-based
    
    try:
        # Get Date (Row 1, same column)
        date_cell = sheet.cell(row=1, column=cell.column)
        date_val = date_cell.value
        
        # Get Time (Column 1, same row)
        time_cell = sheet.cell(row=cell.row, column=1)
        time_val = time_cell.value
        
        if isinstance(date_val, datetime):
            date_part = date_val.date()
        else:
            # Try parsing string? skipping for now
            return None
            
        if isinstance(time_val, time):
            time_part = time_val
        elif isinstance(time_val, datetime):
            time_part = time_val.time()
        else:
            return None
            
        start_dt = datetime.combine(date_part, time_part)
        end_dt = start_dt + timedelta(minutes=config.EVENT_DURATION_MINUTES)
        
        return {
            'start': start_dt,
            'end': end_dt,
            'title': "PHYS2002 Focus Session"
        }
    except Exception:
        return None
